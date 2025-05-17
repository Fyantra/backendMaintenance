#export pour PDF 
from reportlab.lib.pagesizes import legal
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from django.http import HttpResponse
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch

from django.apps import apps
from datetime import datetime
from django.utils import timezone
import os
import urllib.parse
from django.conf import settings
from django.db.models.fields.files import ImageFieldFile, ImageField
from django.db.models import DateField, DateTimeField
from django.shortcuts import get_object_or_404
#export pour excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
#export pour HTML/PDF
from django.template.loader import render_to_string
from weasyprint import HTML
from django.http import HttpResponse
from ..models import Tache, ActiviteTache, ActiviteTachePieceDetachee
#export CSV
import csv

class BaseExporter:
    model_name = None
    title = ''
    columns = {}  # Format: {'field_name': 'Libellé personnalisé'}
    excluded_fields = ['deleted_at']  # Champs à exclure par défaut
    date_format = "%d/%m/%Y"
    datetime_format = "%d/%m/%Y à %H:%M"
    pagesize = legal
    
    @classmethod
    def get_model(cls):
        try:
            return apps.get_model('maintenance', cls.model_name)
        except LookupError:
            raise ValueError(f"Modèle {cls.model_name} non trouvé")

    @classmethod
    def get_field_format(cls, field_name):
        """Détermine le format approprié pour un champ donné"""
        field = cls.get_model()._meta.get_field(field_name)
        if isinstance(field,DateTimeField):
            return cls.datetime_format
        elif isinstance(field, DateField):
            return cls.date_format
        return None

    @classmethod
    def format_value(cls, field_name, value):
        """Formate la valeur selon son type"""
        if value is None:
            return ""
            
        # Gestion spéciale pour les images (uniquement pour PDF)
        if isinstance(value, ImageFieldFile):
            return value  # Le PDF gérera le rendu de l'image
        
        # Formatage des dates
        field_format = cls.get_field_format(field_name)
        if field_format and hasattr(value, 'strftime'):
            return value.strftime(field_format)
            
        return str(value)
    
    @classmethod
    def get_export_columns(cls, format_type):
        """Retourne les colonnes à exporter selon le format"""
        if format_type == 'pdf':
            return cls.columns
        return {
            field: label 
            for field, label in cls.columns.items()
            if not cls.is_image_field(field)
        }
    
    @classmethod
    def is_image_field(cls, field_name):
        """Détecte si un champ est de type ImageField"""
        try:
            field = cls.get_model()._meta.get_field(field_name)
            return isinstance(field, ImageField)
        except:
            return False
    
    @classmethod
    def get_export_data(cls, format_type):
        """Prépare les données selon le format"""
        try:
            model = cls.get_model()
            columns = cls.get_export_columns(format_type)
            
            queryset = model.objects.filter(deleted_at__isnull=True).order_by('-date_creation')
            headers = list(columns.values())
            
            data = []
            for obj in queryset:
                row = []
                for field_name in columns.keys():
                    try:
                        value = getattr(obj, field_name)
                        if format_type == 'pdf' and cls.is_image_field(field_name):
                            row.append(value)
                        else:
                            row.append(cls.format_value(field_name, value))
                    except Exception as e:
                        row.append(f"Erreur: {str(e)}")
                
                data.append(row)
            
            return headers, data
            
        except Exception as e:
            raise ValueError(f"Erreur préparation données {format_type}: {str(e)}")

    @classmethod
    def export_pdf(cls, request):
        """Export PDF avec mise en page avancée"""
        try:
            headers, data = cls.get_export_data('pdf')
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{cls.model_name}.pdf"'
            
            doc = SimpleDocTemplate(response, pagesize=cls.pagesize,topMargin=50, bottomMargin=50)
            elements = []
            
            # En-tête avec logo
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo/logo.png')
            if os.path.exists(logo_path):
                elements.append(Image(logo_path, width=80, height=80, hAlign='RIGHT'))
            
            # Titre et date
            styles = getSampleStyleSheet()
            elements.extend([
                Paragraph(f"Date de génération : {datetime.now().strftime(cls.datetime_format)}", styles['Normal']),
                Spacer(1, 12),
                Paragraph(f"Liste des {cls.title}", styles['Title']),
                Spacer(1, 12)
            ])
            
            # Préparation des données pour PDF
            pdf_data = [headers]
            for row in data:
                pdf_row = []
                for i, value in enumerate(row):
                    field_name = list(cls.columns.keys())[i]
                    if isinstance(value, ImageFieldFile):
                        img_path = os.path.join(settings.MEDIA_ROOT, value.name)
                        pdf_row.append(Image(img_path, width=50, height=50) if os.path.exists(img_path) else "Image")
                    else:
                        pdf_row.append(cls.format_value(field_name, value))
                pdf_data.append(pdf_row)
            
            # Création du tableau PDF
            table = Table(pdf_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.white), # En-têtes
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0b1e48')), # Texte des en-têtes
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), # Centrer horizontalement tout le texte
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Centrer verticalement tout le texte
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Police en gras pour les en-têtes
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), # Grille gris
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(table)
            
            elements.extend([
                Spacer(1, 12),
                Paragraph(f"Nombre {cls.title}: {len(data)}", styles['Normal']),  
            ])
            
            doc.build(elements)
            return response
            
        except Exception as e:
            return HttpResponse(f"Erreur lors de l`export PDF: {str(e)}", status=500)
        
    @classmethod
    def export_csv(cls, request):
        try:
            cls.datetime_format = "%d/%m/%Y %H:%M"
            headers, data = cls.get_export_data('csv')
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{cls.model_name}.csv"'
            
            writer = csv.writer(response)  
            writer.writerow(headers)
            writer.writerows(data)
            
            return response
            
        except Exception as e:
            return HttpResponse(f"Erreur dans l`export CSV: {str(e)}", status=500)
        
    @classmethod
    def export_excel(cls, request):
        """Export excel avec mis en page des colonnes"""
        try:
            cls.datetime_format = "%d/%m/%Y %H:%M"
            headers, data = cls.get_export_data('excel')
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{cls.model_name}.xlsx"'
            
            wb = Workbook()
            ws = wb.active
            ws.title = cls.model_name[:30]  # Excel limite à 31 caractères
            
            # Style des en-têtes
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='9cd084', end_color='9cd084', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Données
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value).alignment = Alignment(vertical='center')
            
            # Ajustement automatique des colonnes
            for column in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(response)
            return response
            
        except Exception as e:
            return HttpResponse(f"Erreur lors de l`export Excel: {str(e)}", status=500)
        
#-----------------------------------------------------------------------------------------------------------------------------
class TacheExporter:
    @classmethod
    def get_safe_file_url(cls, path):
        """
        Transforme un chemin de fichier local Windows en URL utilisable par WeasyPrint,
        en gardant la lettre du lecteur intacte (ex: 'D:') et en encodant proprement le reste.
        """
        drive, rest = os.path.splitdrive(path)
        safe_rest = urllib.parse.quote(rest.replace('\\', '/'))
        return f"file:///{drive}{safe_rest}"
    
    @classmethod
    def export_pdf(cls, request, pk=None):
        if pk:
            # Export d’une seule tâche : vue détaillée
            tache = get_object_or_404(Tache, pk=pk, deleted_at__isnull=True)

            activites = ActiviteTache.objects.filter(tache=tache, deleted_at__isnull=True)

            total_heures = sum(a.temps_passe_heure or 0 for a in activites)
            total_minutes = sum(a.temps_passe_minute or 0 for a in activites)
            total_heures += total_minutes // 60
            total_minutes %= 60
            total_duree_tache = f"{total_heures}h {total_minutes}mn"

            if tache.machine and tache.machine.image:
                image_abs_path = os.path.join(settings.MEDIA_ROOT, tache.machine.image.name)
                image_path = cls.get_safe_file_url(image_abs_path)
            else:
                image_path = None
                
            cout_total_pieces = 0
            activites_data = []
            for activite in activites:
                pieces = ActiviteTachePieceDetachee.objects.filter(activite_tache=activite)
                pieces_data = []
                for piece in pieces:
                    total = (piece.quantite or 0) * (piece.prix_piece_detachees or 0)
                    cout_total_pieces += total
                    pieces_data.append({
                        'instance': piece,
                        'total': total,
                    })
                activites_data.append({
                    'instance': activite,
                    'pieces': pieces_data,
                })

            html_string = render_to_string('export_tache_detail.html', {
                'tache': tache,
                'activites': activites_data,
                'total_duree_tache': total_duree_tache,
                'total_activites': len(activites_data),
                'now': timezone.now(),
                'image_path': image_path,
                'cout_total_pieces': cout_total_pieces, 
            })

        else:
            # Export de toutes les tâches : vue liste
            taches = Tache.objects.filter(deleted_at__isnull=True).order_by('-date_creation')

            taches_data = []
            total_activites = 0

            for tache in taches:
                activites = ActiviteTache.objects.filter(tache=tache, deleted_at__isnull=True)
                total_activites += activites.count()

                total_heures = sum(a.temps_passe_heure or 0 for a in activites)
                total_minutes = sum(a.temps_passe_minute or 0 for a in activites)
                total_heures += total_minutes // 60
                total_minutes %= 60
                total_duree_tache = f"{total_heures}h {total_minutes}mn"

                activites_data = []
                for activite in activites:
                    pieces = ActiviteTachePieceDetachee.objects.filter(activite_tache=activite)
                    pieces_data = []
                    for piece in pieces:
                        total = (piece.quantite or 0) * (piece.prix_piece_detachees or 0)
                        pieces_data.append({
                            'instance': piece,
                            'total': total,
                        })

                    activites_data.append({
                        'instance': activite,
                        'pieces': pieces_data,
                    })

                taches_data.append({
                    'instance': tache,
                    'activites': activites_data,
                    'total_duree_tache': total_duree_tache,
                })

            html_string = render_to_string('export_taches.html', {
                'taches_data': taches_data,
                'total_taches': len(taches_data),
                'total_activites': total_activites,
                'now': timezone.now(),
            })

        pdf_file = HTML(string=html_string).write_pdf()

        filename = f"export_taches_ID_{pk}.pdf" if pk else "export_taches.pdf"
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @classmethod
    def export_excel(cls, request):

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="export_taches.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Tâches de Maintenance"

        # Styles
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        subheader_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        normal_font = Font(size=11)

        # En-têtes des colonnes pour les tâches
        tache_headers = [
            'Description du tâche', 'Équipement', 'Tâche créée le',
            'Programmé pour le', 'Temps de maintenance planifiée',
            'Degré', 'Temps d\'arrêt planifiée'
        ]
        ws.append(tache_headers)

        # Appliquer le style aux en-têtes
        for col in range(1, len(tache_headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        taches = Tache.objects.filter(deleted_at__isnull=True).order_by('-date_creation')
        row_num = 2

        for tache in taches:
            # Données de la tâche
            ws.append([
                tache.description,
                tache.machine.nom_machine if tache.machine else '',
                tache.date_creation.strftime('%d/%m/%Y à %H:%M') if tache.date_creation else '',
                f"{tache.date_debut.strftime('%d/%m/%Y %H:%M') if tache.date_debut else ''} - {tache.date_fin.strftime('%d/%m/%Y %H:%M') if tache.date_fin else ''}",
                f"{tache.temps_maintenance_heure or 0}h {tache.temps_maintenance_minute or 0}mn",
                tache.motif_tache.nom_motif_tache if tache.motif_tache else '',
                f"{tache.temps_arret_heure or 0}h {tache.temps_arret_minute or 0}mn"
            ])

            # Style des données de la tâche
            for col in range(1, len(tache_headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

            row_num += 1

            # Activités de la tâche
            activites = ActiviteTache.objects.filter(tache=tache, deleted_at__isnull=True)
            if activites.exists():
                # En-têtes des activités
                activite_headers = [
                    '',  
                    'Activité réalisée le', 'Description de l\'activité',
                    'Pièces détachées utilisées', 'Temps passé', 'Activité créée le'
                ]
                ws.append(activite_headers)

                # Style des en-têtes d'activités
                for col in range(1, len(activite_headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if col > 1:  # Ne pas styliser la première cellule vide
                        cell.fill = subheader_fill
                        cell.font = bold_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left')

                row_num += 1

                for activite in activites:
                    pieces = ActiviteTachePieceDetachee.objects.filter(activite_tache=activite)
                    pieces_text = "\n".join(
                        f"• {piece.pieces_detachees.nom_piecedetache} : {piece.quantite} * {piece.prix_piece_detachees} Ar"
                        for piece in pieces
                    )

                    ws.append([
                        '',  
                        activite.date_realisation.strftime('%d/%m/%Y %H.%M') if activite.date_realisation else '',
                        activite.description,
                        pieces_text,
                        f"{activite.temps_passe_heure or 0}h {activite.temps_passe_minute or 0}mn",
                        activite.date_creation.strftime('%d/%m/%Y à %H:%M') if activite.date_creation else ''
                    ])

                    # Style des données d'activité
                    for col in range(1, len(activite_headers) + 1):
                        cell = ws.cell(row=row_num, column=col)
                        cell.font = normal_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    row_num += 1

                # Ajouter une ligne vide entre les tâches
                row_num += 1

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(response)
        return response
    
    @classmethod
    def export_csv(cls, request):

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export_taches.csv"'

        writer = csv.writer(response)
        
        # En-têtes des colonnes pour les tâches
        writer.writerow([
            'Type', 'Description du tâche', 'Équipement', 'Tâche créée le',
            'Programmé pour le', 'Temps de maintenance planifiée',
            'Degré', 'Temps d\'arrêt planifiée'
        ])

        taches = Tache.objects.filter(deleted_at__isnull=True).order_by('-date_creation')

        for tache in taches:
            # Données de la tâche
            writer.writerow([
                'TÂCHE',
                tache.description,
                tache.machine.nom_machine if tache.machine else '',
                tache.date_creation.strftime('%d/%m/%Y %H:%M') if tache.date_creation else '',
                f"{tache.date_debut.strftime('%d/%m/%Y %H:%M') if tache.date_debut else ''} - {tache.date_fin.strftime('%d/%m/%Y %H:%M') if tache.date_fin else ''}",
                f"{tache.temps_maintenance_heure or 0}h {tache.temps_maintenance_minute or 0}mn",
                tache.motif_tache.nom_motif_tache if tache.motif_tache else '',
                f"{tache.temps_arret_heure or 0}h {tache.temps_arret_minute or 0}mn"
            ])

            # Activités de la tâche
            activites = ActiviteTache.objects.filter(tache=tache, deleted_at__isnull=True)
            for activite in activites:
                pieces = ActiviteTachePieceDetachee.objects.filter(activite_tache=activite)
                pieces_text = " | ".join(
                    f"• {piece.pieces_detachees.nom_piecedetache} : {piece.quantite} * {piece.prix_piece_detachees} Ar"
                    for piece in pieces
                )

                writer.writerow([
                    'ACTIVITÉ','','', 
                    activite.date_realisation.strftime('%d/%m/%Y %H.%M') if activite.date_realisation else '',
                    '','',  activite.description,'',pieces_text,
                    f"{activite.temps_passe_heure or 0}h {activite.temps_passe_minute or 0}mn",
                    activite.date_creation.strftime('%d/%m/%Y %H:%M') if activite.date_creation else ''
                ])

            # Ligne vide entre les tâches
            writer.writerow([])

        return response
